from typing import Any, Callable, Generic, List, Optional, TypeVar

from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from folio_upm.utils import log_factory
from folio_upm.utils.upm_env import Env
from folio_upm.xlsx import ws_constants

T = TypeVar("T")


class Column(Generic[T]):
    def __init__(self, n: str, w: float, f: Callable[[T], Optional[Any]]):
        self.n = n
        self.w = w
        self.f = f


class AbstractWorksheet(Generic[T]):

    def __init__(self, ws: Worksheet, title: str, data: Any, columns: List[Column[T]]):
        self._log = log_factory.get_logger(self.__class__.__name__)
        self._ws = ws
        self._data = data
        self._ws.title = title
        self._columns = columns
        self._row_num = 1

    def fill(self):
        self._log.debug("Worksheet fill initialized: '%s'...", self._ws.title)
        self._populate_headers()
        self._fill_rows()
        self._ws.freeze_panes = "A2"
        self._ws.auto_filter.ref = self._ws.dimensions
        self._log.info("Worksheet filled with data: '%s', rows added: %s", self._ws.title, self._row_num - 1)

    def _populate_headers(self):
        for col_num, header in enumerate(self._columns, start=1):
            cell = self._ws.cell(column=col_num, row=self._row_num, value=header.n)
            cell.font = ws_constants.header_font
            cell.border = ws_constants.thin_border
            cell.alignment = ws_constants.header_cell_alignment
            cell.fill = ws_constants.light_gray_fill
            self._ws.column_dimensions[cell.column_letter].width = header.w  # type: ignore
            self._ws.row_dimensions[self._row_num].height = 20
        self._row_num += 1

    def _fill_rows(self):
        for value in self._get_iterable_data():
            colors_enabled = Env().get_bool_cached("ENABLE_REPORT_COLORING", default_value=False)
            _fill = self._get_row_fill_color(value) if colors_enabled else ws_constants.almost_white_fill
            self._add_row(self._map_value(value), fill=_fill)

    def title(self):
        return self._ws.title

    def _map_value(self, value: T) -> List[Optional[Any]]:
        return [col.f(value) for col in self._columns]

    def _get_iterable_data(self) -> List[T]:
        return self._data

    def _add_row(self, row: List[Optional[Any]], fill: PatternFill):
        for col_num, cell_value in enumerate(row, start=1):
            cell = self._ws.cell(column=col_num, row=self._row_num, value=cell_value)
            cell.font = ws_constants.data_font
            cell.border = ws_constants.thin_border
            cell.alignment = ws_constants.data_cell_alignment
            if fill:
                cell.fill = fill
        self._row_num += 1

    def _get_row_fill_color(self, value: T) -> PatternFill:
        return ws_constants.almost_white_fill
