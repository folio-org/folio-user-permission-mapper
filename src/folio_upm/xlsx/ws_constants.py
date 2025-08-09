from openpyxl.descriptors.base import Float
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

almost_white = "fafafa"
light_red = "ffe1e5"
light_gray = "bababa"
lightest_gray = "d9d9d9"
light_green = "c6efe3"
light_yellow = "ffffcc"

almost_white_fill = PatternFill(start_color=almost_white, end_color=almost_white, fill_type="darkHorizontal")
light_gray_fill = PatternFill(start_color=lightest_gray, end_color=lightest_gray, fill_type="darkHorizontal")
light_green_fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="darkHorizontal")
light_red_fill = PatternFill(start_color=light_red, end_color=light_red, fill_type="darkHorizontal")
light_yellow_fill = PatternFill(start_color=light_yellow, end_color=light_yellow, fill_type="darkHorizontal")

thin_border = Border(
    left=Side(style="thin", color=light_gray),
    right=Side(style="thin", color=light_gray),
    top=Side(style="thin", color=light_gray),
    bottom=Side(style="thin", color=light_gray),
)

data_font = Font(name="Consolas", bold=False, italic=False, size=10)
header_font = Font(name="Consolas", bold=True, size=11)
header_cell_alignment = Alignment(horizontal="left", vertical="center")
data_cell_alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

uuid_cw = Float("40")
num_short_cw = Float("25")
num_long_cw = Float("40")
bool_cw_short = Float("20")
bool_long_cw = Float("30")
number_cw_long = Float("40")

type_cw = Float("25")
note_cw = Float("35")
desc_short_cw = Float("60")
desc_med_cw = Float("90")
desc_long_cw = Float("120")

role_id_cw = Float("60")
role_name_cw = Float("80")
ps_name_cw = Float("80")
