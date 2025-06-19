import io
from collections import OrderedDict
from typing import List
from typing import OrderedDict as OrdDict

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from folio_upm.dto.models import AnalysisResult
from folio_upm.utils import log_factory

_log = log_factory.get_logger(__name__)
_light_green_fill = PatternFill(start_color="c6efe3", end_color="c6efe3", fill_type="darkHorizontal")
_light_red_fill = PatternFill(start_color="ffe1e5", end_color="ffe1e5", fill_type="darkHorizontal")
_thin_border = Border(
    left=Side(style="thin", color="bababa"),
    right=Side(style="thin", color="bababa"),
    top=Side(style="thin", color="bababa"),
    bottom=Side(style="thin", color="bababa"),
)


def generate_report(analysis_result: AnalysisResult):
    _log.info("generating XLSX report...")
    wb = Workbook()
    wb.remove(wb.active)

    __fill_user_stats(analysis_result, wb.create_sheet())
    __fill_mutable_pss(analysis_result, wb.create_sheet())
    __fill_all_pss(analysis_result, wb.create_sheet())
    __fill_permission_set_nesting_ws(analysis_result, wb.create_sheet())
    __fill_users_mutable_pss_ws(analysis_result, wb.create_sheet())
    __fill_users_system_pss_ws(analysis_result, wb.create_sheet())
    __fill_permission_permission_sets_ws(analysis_result, wb.create_sheet())
    __fill_roles(analysis_result, wb.create_sheet())
    __fill_role_users(analysis_result, wb.create_sheet())
    __fill_role_capabilities(analysis_result, wb.create_sheet())

    return __to_byte_doc(wb)


def __fill_user_stats(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "User Stats", ["User Id", "# of Mutable PS", "# of System PS", "# of unmatched PS", "# of Total PS"])
    for user_id, uph in analysis_result.usersPermissionSets.items():
        mutable_permission_num = len(uph.mutablePermissions)
        system_permission_num = len(uph.systemPermissions)
        unmatched_permission_num = len([x for x in uph.systemPermissions if not x[1]])
        row = [
            user_id,
            mutable_permission_num,
            system_permission_num,
            unmatched_permission_num,
            mutable_permission_num + system_permission_num,
        ]
        ws.append(row)

    formatting_pattern = f"A2:E{ws.max_row}"
    ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=["$D2=0"], fill=_light_green_fill))
    ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=["$D2>0"], fill=_light_red_fill))
    __format_worksheet(ws, [("A", 45), ("B", 18), ("C", 18), ("D", 18), ("E", 18)])


def __fill_mutable_pss(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Mutable PS", ["Name", "Display Name", "# of PS", "# of Flat PS", "Description"])

    permissions = analysis_result.permissionSets.items()
    for permission_name, permission in permissions:
        if permission.mutable:
            sub_ps_num = len(permission.subPermissions)
            assigned_flat_ps_num = len(analysis_result.flatPermissionSets.get(permission_name, []))
            row = [
                permission_name,
                permission.displayName,
                sub_ps_num,
                assigned_flat_ps_num,
                permission.description,
            ]
            ws.append(row)

    __format_worksheet(ws, [("A", 45), ("B", 80), ("C", 14), ("D", 14), ("E", 120)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_all_pss(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "All PS", ["Name", "Display Name", "Mutable", "Description"])

    permissions = analysis_result.permissionSets.items()
    for permission_name, permission in permissions:
        mutable = str(permission.mutable).lower()
        ws.append([permission_name, permission.displayName, mutable, permission.description])

    __format_worksheet(ws, [("A", 45), ("B", 80), ("C", 18), ("D", 120)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_permission_set_nesting_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "PermissionSets Nesting", ["PS", "PS Parent"])

    nested_perms_items = analysis_result.permissionSetsNesting.items()
    for permission_name, parent_permission_names in nested_perms_items:
        if not parent_permission_names:
            ws.append([permission_name, "null"])
        else:
            for perm in parent_permission_names:
                ws.append([permission_name, perm])

    __format_worksheet(ws, [("A", 45), ("B", 45)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_users_mutable_pss_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Users-Mutable PS", ["User UUID", "PS"])
    mutable_permissions = OrderedDict()
    for user_id, up_holder in analysis_result.usersPermissionSets.items():
        mutable_permissions[user_id] = up_holder.mutablePermissions
    __fill_map_values(mutable_permissions, ws, fill_nulls=True)
    ws.conditional_formatting.add(f"A2:B{ws.max_row}", FormulaRule(formula=['$B2="null"'], fill=_light_red_fill))


def __fill_users_system_pss_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Users-System PS", ["User UUID", "PS", "In Mutable PS", "Defined"])
    mutable_permissions = OrderedDict()
    for user_id, up_holder in analysis_result.usersPermissionSets.items():
        mutable_permissions[user_id] = up_holder.systemPermissions
    __fill_map_values(mutable_permissions, ws, fill_nulls=True, map_func=lambda x: [x[0], __bts(x[1]), __bts(x[2])])

    formatting_pattern = f"A2:D{ws.max_row}"
    formula_green = 'AND($C2="true",$D2="true")'
    formula_red = 'OR($C2="false",$D2="false")'
    ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=[formula_green], fill=_light_green_fill))
    ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=[formula_red], fill=_light_red_fill))

    __format_worksheet(ws, [("A", 45), ("B", 80), ("C", 18), ("D", 18)])


def __fill_permission_permission_sets_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "PermissionSet-PermissionSets", ["PS", "Child PS"])

    items = analysis_result.permissionPermissionSets.items()
    for permission_name, sub_permission_names in items:
        if not sub_permission_names:
            ws.append([permission_name, "null"])
        else:
            for sub_permission in sub_permission_names:
                ws.append([permission_name, sub_permission])
    __format_worksheet(ws, [("A", 80), ("B", 80)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_roles(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Roles", ["ID", "Source PS", "Name", "Description"])
    for role in analysis_result.roles:
        ws.append([role.id, role.source, role.name, role.description])
    __format_worksheet(ws, [("A", 50), ("B", 50), ("C", 65), ("D", 120)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_role_users(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Role-Users", ["Role ID", "User UUID"])

    __fill_map_values(analysis_result.roleUsers, ws, fill_nulls=False)
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_role_capabilities(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Role-Capabilities", ["Role ID", "PS", "Type"])
    for role_id, holder in analysis_result.roleCapabilities.items():
        for capabilitySetPS in holder.capabilitySetPS:
            ws.append([role_id, capabilitySetPS, "capabilitySet"])
        for capabilityPS in holder.capabilityPS:
            ws.append([role_id, capabilityPS, "capability"])

    __format_worksheet(ws, [("A", 50), ("B", 80), ("C", 30)])
    _log.info(f"Worksheet finished: {ws.title}")


def __init_ws(ws, title: str, headers: List[str] = None):
    ws.title = title
    _log.info(f"Filling worksheet: {ws.title}...")
    if headers:
        ws.append(headers)


def __fill_map_values(values_map: OrdDict[str, List[str]], ws, fill_nulls: bool = True, map_func=lambda x: [x]):
    role_users = values_map.items()
    for key, values_list in role_users:
        if not values_list and fill_nulls:
            ws.append([key, "null", None])
            continue
        for value in values_list:
            ws.append([key, *map_func(value)])
    __format_worksheet(ws, [("A", 45), ("B", 45)])
    _log.info(f"Worksheet finished: {ws.title}")


def __get_mutable_permission_names(analysis_json):
    mutable_permissions = analysis_json.get("allPermissions", [])
    return set([x.get("permissionName") for x in mutable_permissions if x.get("mutable", False)])


def __format_worksheet(ws: Worksheet, dimensions):
    for dimension in dimensions:
        columnName = dimension[0]
        columnWidth = dimension[1]
        ws.column_dimensions[columnName].width = columnWidth

    cells_font = Font(name="Consolas", bold=False, italic=False, size=11)
    header_font = Font(name="Consolas", bold=True, size=11)
    __apply_font_for_cells(ws, f"A1:E{ws.max_row}", cells_font)
    __apply_font_for_cells(ws, "A1:E1", header_font)
    __apply_borders(ws, len(dimensions))


def __apply_borders(ws: Worksheet, end_col):

    # Apply the border to each cell in the range
    for row in range(1, ws.max_row + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin_border


def __to_byte_doc(wb):
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def __bts(value: bool) -> str:
    return str(value).lower()


def __apply_font_for_cells(ws, key, font):
    """
    Apply the given font to all cells in the worksheet.
    """
    for row in ws[key]:
        for cell in row:
            cell.font = font
