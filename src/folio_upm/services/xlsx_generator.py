import io
from collections import OrderedDict
from typing import List
from typing import OrderedDict as OrdDict

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from folio_upm.dto.results import AnalysisResult
from folio_upm.utils import log_factory

_log = log_factory.get_logger(__name__)
_light_green_fill = PatternFill(start_color="c6efe3", end_color="c6efe3", fill_type="darkHorizontal")
_light_red_fill = PatternFill(start_color="ffe1e5", end_color="ffe1e5", fill_type="darkHorizontal")
_thin_border = Border(
    left=Side(style="thin", color="bababa"),
    right=Side(style="thin", color="bababa"),
    top=Side(style="thin", color="bababa"),
    bottom=Side(style="thin", color="bababa"),
)

_cells_font = Font(name="Consolas", bold=False, italic=False, size=11)
_header_font = Font(name="Consolas", bold=True, size=11)


class XlsxReportGenerator:
    """
    Generates XLSX report from the analysis result.
    """

    def __init__(self, analysis_result: AnalysisResult):
        self._analysis_result = analysis_result
        self._wb = Workbook()
        self._wb.remove(self._wb.active)
        self.__fill_worksheet()
        self._report_bytes = self.__to_byte_doc()

    def get_report_bytes(self) -> io.BytesIO:
        return self._report_bytes

    def __fill_worksheet(self):
        _log.info("generating XLSX report...")

        self.__fill_user_stats()
        self.__fill_mutable_pss()
        self.__fill_all_pss()
        self.__fill_permission_set_nesting_ws()
        self.__fill_users_mutable_pss_ws()
        self.__fill_users_system_pss_ws()
        self.__fill_permission_permission_sets_ws()
        self.__fill_roles()
        self.__fill_role_users()
        self.__fill_role_capabilities()

    def __fill_user_stats(self):
        headers = ["User Id", "# of Mutable PS", "# of System PS", "# of unmatched PS", "# of Total PS"]
        ws = _Utils.init_ws(self._wb, "User Stats", headers)
        for user_id, uph in self._analysis_result.usersPermissionSets.items():
            mutable_permission_num = len(uph.mutablePermissions)
            system_permission_num = len(uph.systemPermissions)
            unmatched_permission_num = len([x for x in uph.systemPermissions if not x[1]])
            row = [
                user_id,
                mutable_permission_num,
                system_permission_num,
                unmatched_permission_num,
                mutable_permission_num + system_permission_num,
            ]
            ws.append(row)

        formatting_pattern = f"A2:E{ws.max_row}"
        ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=["$D2=0"], fill=_light_green_fill))
        ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=["$D2>0"], fill=_light_red_fill))
        _Utils.format_worksheet(ws, [("A", 45), ("B", 18), ("C", 18), ("D", 18), ("E", 18)])

    def __fill_mutable_pss(self):
        headers = ["Name", "Display Name", "# of PS", "# of Flat PS", "Description"]
        ws = _Utils.init_ws(self._wb, "Mutable PS", headers)
        analysis_result = self._analysis_result.permsAnalysisResult
        permissions = analysis_result.permissionSets.items()

        for permission_name, permission in permissions:
            if permission.mutable:
                sub_ps_num = len(permission.subPermissions)
                assigned_flat_ps_num = len(analysis_result.flatPermissionSets.get(permission_name, []))
                row = [
                    permission_name,
                    permission.displayName,
                    sub_ps_num,
                    assigned_flat_ps_num,
                    permission.description,
                ]
                ws.append(row)

        _Utils.format_worksheet(ws, [("A", 45), ("B", 80), ("C", 14), ("D", 14), ("E", 120)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_all_pss(self):
        ws = _Utils.init_ws(self._wb, "All PS", ["Name", "Display Name", "Mutable", "Description"])
        analysis_result = self._analysis_result
        permissions = analysis_result.mutablePermissionSets.items()
        for permission_name, permission in permissions:
            mutable = str(permission.mutable).lower()
            ws.append([permission_name, permission.displayName, mutable, permission.description])

        _Utils.format_worksheet(ws, [("A", 45), ("B", 80), ("C", 18), ("D", 120)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_permission_set_nesting_ws(self):
        ws = _Utils.init_ws(self._wb, "PermissionSets Nesting", ["PS", "PS Parent"])

        nested_perms_items = self._analysis_result.permissionSetsNesting.items()
        for permission_name, parent_permission_names in nested_perms_items:
            if not parent_permission_names:
                ws.append([permission_name, "null"])
            else:
                for perm in parent_permission_names:
                    ws.append([permission_name, perm])

        _Utils.format_worksheet(ws, [("A", 45), ("B", 45)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_users_mutable_pss_ws(self):
        ws = _Utils.init_ws(self._wb, "Users-Mutable PS", ["User UUID", "PS"])
        mutable_permissions = OrderedDict()
        for user_id, up_holder in self._analysis_result.usersPermissionSets.items():
            mutable_permissions[user_id] = up_holder.mutablePermissions
        self.__fill_map_values(mutable_permissions, ws, fill_nulls=True)
        ws.conditional_formatting.add(f"A2:B{ws.max_row}", FormulaRule(formula=['$B2="null"'], fill=_light_red_fill))

    def __fill_users_system_pss_ws(self):
        ws = _Utils.init_ws(self._wb, "Users-System PS", ["User UUID", "PS", "In Mutable PS", "Defined"])
        mutable_permissions = OrderedDict()
        for user_id, up_holder in self._analysis_result.usersPermissionSets.items():
            mutable_permissions[user_id] = up_holder.systemPermissions
        mapper_func = lambda x: [x[0], x[1], x[2]]
        XlsxReportGenerator.__fill_map_values(mutable_permissions, ws, fill_nulls=True, map_func=mapper_func)

        formatting_pattern = f"A2:D{ws.max_row}"
        formula_green = 'AND($C2="true",$D2="true")'
        formula_red = 'OR($C2="false",$D2="false")'
        ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=[formula_green], fill=_light_green_fill))
        ws.conditional_formatting.add(formatting_pattern, FormulaRule(formula=[formula_red], fill=_light_red_fill))

        _Utils.format_worksheet(ws, [("A", 45), ("B", 80), ("C", 18), ("D", 18)])

    def __fill_permission_permission_sets_ws(self):
        ws = _Utils.init_ws(self._wb, "PermissionSet-PermissionSets", ["PS", "Child PS"])

        items = self._analysis_result.permissionPermissionSets.items()
        for permission_name, sub_permission_names in items:
            if not sub_permission_names:
                ws.append([permission_name, "null"])
            else:
                for sub_permission in sub_permission_names:
                    ws.append([permission_name, sub_permission])
        _Utils.format_worksheet(ws, [("A", 80), ("B", 80)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_roles(self):
        ws = _Utils.init_ws(self._wb, "Roles", ["ID", "Source PS", "Name", "Description"])
        for role in self._analysis_result.roles:
            ws.append([role.id, role.source, role.name, role.description])
        _Utils.format_worksheet(ws, [("A", 50), ("B", 50), ("C", 65), ("D", 120)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_role_users(self):
        ws = _Utils.init_ws(self._wb, "Role-Users", ["Role ID", "User UUID"])

        _Utils.format_worksheet(self._analysis_result.roleUsers, ws, fill_nulls=False)
        _log.info(f"Worksheet finished: {ws.title}")

    def __fill_role_capabilities(self):
        ws = _Utils.init_ws(self._wb, "Role-Capabilities", ["Role ID", "PS", "Type"])
        for role_id, holder in self._analysis_result.roleCapabilities.items():
            for capabilitySetPS in holder.capabilitySetPS:
                ws.append([role_id, capabilitySetPS, "capabilitySet"])
            for capabilityPS in holder.capabilityPS:
                ws.append([role_id, capabilityPS, "capability"])

        _Utils.format_worksheet(ws, [("A", 50), ("B", 80), ("C", 30)])
        _log.info(f"Worksheet finished: {ws.title}")

    def __to_byte_doc(self):
        excel_buffer = io.BytesIO()
        self._wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    @staticmethod
    def __fill_map_values(values_map: OrdDict[str, List[str]], ws, fill_nulls: bool = True, map_func=lambda x: [x]):
        role_users = values_map.items()
        for key, values_list in role_users:
            if not values_list and fill_nulls:
                ws.append([key, "null", None])
                continue
            for value in values_list:
                ws.append([key, *map_func(value)])
        _Utils.format_worksheet(ws, [("A", 45), ("B", 45)])
        _log.info(f"Worksheet finished: {ws.title}")


class _Utils:

    @staticmethod
    def init_ws(wb: Workbook, title: str, headers: List[str] = None) -> Worksheet:
        ws = wb.create_sheet()
        ws.title = title
        _log.info(f"Filling worksheet: {ws.title}...")
        if headers:
            ws.append(headers)
        return ws

    @staticmethod
    def format_worksheet(ws: Worksheet, dimensions):
        for dimension in dimensions:
            columnName = dimension[0]
            columnWidth = dimension[1]
            ws.column_dimensions[columnName].width = columnWidth

        _Utils.__apply_font_for_cells(ws, f"A1:E{ws.max_row}", _cells_font)
        _Utils.__apply_font_for_cells(ws, "A1:E1", _header_font)
        _Utils.__apply_borders(ws, len(dimensions))

    @staticmethod
    def __apply_font_for_cells(ws, key, font):
        """
        Apply the given font to all cells in the worksheet.
        """
        for row in ws[key]:
            for cell in row:
                cell.font = font

    @staticmethod
    def __apply_borders(ws: Worksheet, end_col):
        # Apply the border to each cell in the range
        for row in range(1, ws.max_row + 1):
            for col in range(1, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = _thin_border
