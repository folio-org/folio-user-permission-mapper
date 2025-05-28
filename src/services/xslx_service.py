import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def generate_report(analysis_json):
    permission_users = analysis_json['permissionUsers']

    wb = Workbook()

    __fill_permission_set_netsting_ws(analysis_json, wb.active)
    __fill_users_permission_sets_ws(analysis_json, wb.create_sheet())
    __fill_permission_permission_sets_ws(analysis_json, wb.create_sheet())

    __format_spreadsheet(wb)

    return __to_byte_doc(wb)


def __fill_permission_set_netsting_ws(analysis_json, ws: Worksheet):
    ws.title = "PermissionSets Nesting"
    ws.append(["PS", "PS Parent"])
    mutalbe_permissions = analysis_json['mutablePermissions']
    for permission in mutalbe_permissions:
        permission_name = permission['permissionName']
        parent_permissions = permission.get('childOf', [])
        if not parent_permissions:
            ws.append([permission_name, 'null'])
        else:
            for sub_permission in parent_permissions:
                ws.append([permission_name, sub_permission])


def __fill_users_permission_sets_ws(analysis_json, ws: Worksheet):
    ws.title = "Users-PermissionSets"
    ws.append(["User UUID", "PS"])

    mutable_permission_names = __get_mutable_permission_names(analysis_json)

    permission_users = analysis_json.get('permissionUsers', [])
    for permission_user in permission_users:
        user_id = permission_user.get('userId')
        permissions = permission_user.get('permissions', [])
        for permission in permissions:
            if permission in mutable_permission_names:
                ws.append([user_id, permission])


def __get_mutable_permission_names(analysis_json):
    mutable_permissions = analysis_json.get('mutablePermissions', [])
    return set([x.get('permissionName') for x in mutable_permissions])


def __fill_permission_permission_sets_ws(analysis_json, ws: Worksheet):
    ws.title = "PermissionSet-PermissionSets"
    ws.append(["PS", "Child PS"])
    mutable_permission_names = __get_mutable_permission_names(analysis_json)
    mutalbe_permissions = analysis_json.get('mutablePermissions', [])
    for permission in mutalbe_permissions:
        permission_name = permission['permissionName']
        sub_permissions = permission.get('subPermissions', [])
        filter_func = lambda x: (x in mutable_permission_names and x != permission_name)
        filtered_sub_permissions=list(filter(filter_func, sub_permissions))
        if not filtered_sub_permissions:
            ws.append([permission_name, 'null'])
        else:
            for sub_permission in filtered_sub_permissions:
                ws.append([permission_name, sub_permission])


def __format_spreadsheet(wb: Workbook):
    header_font = Font(bold=True, italic=True)
    ws_names = wb.get_sheet_names()
    for ws_name in ws_names:
        ws = wb.get_sheet_by_name(ws_name)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        applyFontForCells(ws, "A1:B1", header_font)


def __to_byte_doc(wb):
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def applyFontForCells(ws, key, font):
    """
    Apply the given font to all cells in the worksheet.
    """
    for row in ws[key]:
        for cell in row:
            cell.font = font
