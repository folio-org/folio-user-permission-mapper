import io
from collections import OrderedDict
from typing import List
from typing import OrderedDict as OrdDict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from dto.models import AnalysisResult
from utils import log_factory

_log = log_factory.get_logger(__name__)


def generate_report(analysis_result: AnalysisResult):
    _log.info("generating XLSX report...")
    wb = Workbook()

    __fill_permission_set(analysis_result, wb.active)
    __fill_permission_set_nesting_ws(analysis_result, wb.create_sheet())
    __fill_users_permission_sets_ws(analysis_result, wb.create_sheet())
    __fill_users_system_permission_sets_ws(analysis_result, wb.create_sheet())
    __fill_permission_permission_sets_ws(analysis_result, wb.create_sheet())
    __fill_roles(analysis_result, wb.create_sheet())
    __fill_role_users(analysis_result, wb.create_sheet())
    __fill_role_capabilities(analysis_result, wb.create_sheet())
    __fill_role_capability_sets(analysis_result, wb.create_sheet())

    return __to_byte_doc(wb)


def __fill_permission_set(analysis_result: AnalysisResult, ws: Worksheet):
    ws.title = "PS"
    _log.info(f"Filling worksheet: {ws.title}...")
    ws.append(["Name", "Display Name", "Description"])

    permissions = analysis_result.permissionSets.items()
    for permission_name, permission in permissions:
        if permission.mutable:
            ws.append([permission_name, permission.displayName, permission.description])

    __format_worksheet(ws, [("A", 40), ("B", 60), ("C", 100)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_permission_set_nesting_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "PermissionSets Nesting", ["PS", "PS Parent"])

    nested_perms_items = analysis_result.permissionSetsNesting.items()
    for permission_name, parent_permission_names in nested_perms_items:
        if not parent_permission_names:
            ws.append([permission_name, "null"])
        else:
            for perm in parent_permission_names:
                ws.append([permission_name, perm])

    __format_worksheet(ws, [("A", 40), ("B", 40)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_users_permission_sets_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Users-PermissionSets", ["User UUID", "PS"])
    mutable_permissions = OrderedDict()
    for user_id, up_holder in analysis_result.usersPermissionSets.items():
        mutable_permissions[user_id] = up_holder.mutablePermissions
    __fill_map_values(mutable_permissions, ws, fill_nulls=True)


def __fill_users_system_permission_sets_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Users-SystemPermissionSets", ["User UUID", "PS"])
    mutable_permissions = OrderedDict()
    for user_id, up_holder in analysis_result.usersPermissionSets.items():
        mutable_permissions[user_id] = up_holder.systemPermissions
    __fill_map_values(mutable_permissions, ws, fill_nulls=True)


def __fill_permission_permission_sets_ws(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "PermissionSet-PermissionSets", ["PS", "Child PS"])

    items = analysis_result.permissionPermissionSets.items()
    for permission_name, sub_permission_names in items:
        if not sub_permission_names:
            ws.append([permission_name, "null"])
        else:
            for sub_permission in sub_permission_names:
                ws.append([permission_name, sub_permission])
    __format_worksheet(ws, [("A", 80), ("B", 80)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_roles(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Roles", ["ID", "Source PS", "Name", "Description"])
    for role in analysis_result.roles:
        ws.append([role.id, role.source, role.name, role.description])
    __format_worksheet(ws, [("A", 50), ("B", 50), ("C", 50), ("D", 100)])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_role_users(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Role-Users", ["Role ID", "User UUID"])

    __fill_map_values(analysis_result.roleUsers, ws, fill_nulls=False)
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_role_capabilities(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Role-Capabilities", ["Role ID", "PS"])
    _log.info(f"Worksheet finished: {ws.title}")


def __fill_role_capability_sets(analysis_result: AnalysisResult, ws: Worksheet):
    __init_ws(ws, "Role-CapabilitySets", ["Role ID", "PS"])
    _log.info(f"Worksheet finished: {ws.title}")


def __init_ws(ws, title: str, headers: List[str] = None):
    ws.title = title
    _log.info(f"Filling worksheet: {ws.title}...")
    if headers:
        ws.append(headers)


def __fill_map_values(values_map: OrdDict[str, List[str]], ws, fill_nulls: bool = True):
    role_users = values_map.items()
    for key, values_list in role_users:
        if not values_list and fill_nulls:
            ws.append([key, "null"])
            continue
        for value in values_list:
            ws.append([key, value])
    __format_worksheet(ws, [("A", 40), ("B", 40)])
    _log.info(f"Worksheet finished: {ws.title}")


def __get_mutable_permission_names(analysis_json):
    mutable_permissions = analysis_json.get("allPermissions", [])
    return set([x.get("permissionName") for x in mutable_permissions if x.get("mutable", False)])


def __format_worksheet(ws: Worksheet, dimensions):
    for dimension in dimensions:
        columnName = dimension[0]
        columnWidth = dimension[1]
        ws.column_dimensions[columnName].width = columnWidth

    header_font = Font(bold=True, italic=True)
    __apply_font_for_cells(ws, "A1:F1", header_font)


def __to_byte_doc(wb):
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def __apply_font_for_cells(ws, key, font):
    """
    Apply the given font to all cells in the worksheet.
    """
    for row in ws[key]:
        for cell in row:
            cell.font = font
